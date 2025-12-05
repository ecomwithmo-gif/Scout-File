import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter
import os
from openpyxl.comments import Comment
import colorsys
import threading
import subprocess
import platform
import time
from concurrent.futures import ThreadPoolExecutor
import gc

# Header mapping as per requirements
HEADER_MAP = {
    'Brand': 'Brand',
    'Parent ASIN': 'Parent',
    'ASIN': 'ASIN',
    'Product Codes: UPC': 'UPC',
    'Product Codes: EAN': 'EAN',
    'Product Codes: GTIN': 'GTIN',
    'Imported by Code': 'Imported by Code',
    'Title': 'Title',
    'Color': 'Color',
    'Size': 'Size',
    'Bought in past month': 'Sales Badge',
    'Reviews: Rating Count': 'Rating Count',
    'Reviews: Review Count - Format Specific': 'Rating - Child',
    'Sales Rank: Current': 'Sales Rank',
    'Sales Rank: 30 days avg.': 'Sales Rank 30',
    'Sales Rank: 90 days avg.': 'Sales Rank 90',
    'Sales Rank: 180 days avg.': 'Sales Rank 180',
    'Buy Box 🚚: Current': 'Buy Box',
    'Buy Box 🚚: 30 days avg.': 'Buy Box 30',
    'Buy Box 🚚: 90 days avg.': 'Buy Box 90',
    'Buy Box 🚚: 180 days avg.': 'Buy Box 180',
    'Amazon: 90 days OOS': 'AMZ In Stock %',
    'Buy Box: % Amazon 90 days': 'Buy Box: % Amazon 90 days',
    'Amazon: Availability of the Amazon offer': 'Amazon Availability',
    'Count of retrieved live offers: New, FBA': 'FBA',
    'Count of retrieved live offers: New, FBM': 'FBM',
    'FBA Pick&Pack Fee': 'Pick & Pack',
    'Referral Fee %': 'Referral Fee &',
}

class ExcelFormatterApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Formatter')
        self.root.geometry("1200x780")
        self.root.minsize(980, 640)
        self.root.resizable(True, True)
        self.is_fullscreen = False
        self.root.bind('<F11>', self.toggle_fullscreen)
        self.root.bind('<Escape>', self.exit_fullscreen)
        
        # Set clean light theme
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        
        # Initialize variables
        self.file_path = None
        self.df = None
        self.file2_path = None
        self.df2 = None
        self.processing = False
        self.chunk_size = 1000  # Process data in chunks for better performance
        self.last_dir = os.getcwd()
        
        # Configure root window with white gradient background
        self.root.configure(bg="#f5f7fb")
        
        # Main scrollable container so the UI stays usable on smaller windows
        self.main_container = ctk.CTkScrollableFrame(self.root, fg_color="#f5f7fb", corner_radius=0)
        self.main_container.pack(expand=True, fill="both", padx=0, pady=0)
        
        # Create header
        self.create_header()
        
        # Create main content
        self.create_main_content()
        
        # Create progress tracking
        self.create_progress_section()
        
    def create_header(self):
        """Create elegant header"""
        header_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        header_frame.pack(fill="x", pady=(28, 18), padx=32)
        
        title_label = ctk.CTkLabel(
            header_frame, 
            text="Formatter workspace", 
            font=ctk.CTkFont(family="Inter", size=30, weight="bold"),
            text_color="#0f172a"
        )
        title_label.pack(anchor="w", pady=(0, 6))
        
        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Load your files, set costs, and process without the clutter.",
            font=ctk.CTkFont(family="Inter", size=13),
            text_color="#475569"
        )
        subtitle_label.pack(anchor="w")
        
    def create_main_content(self):
        """Create main content area with elegant layout"""
        content_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        content_frame.pack(expand=True, fill="both", padx=32, pady=(0, 24))
        
        # Create two-column layout - Files on left, Settings & Actions on right
        left_panel = ctk.CTkFrame(content_frame, fg_color="transparent")
        left_panel.pack(side="left", expand=True, fill="both", padx=(0, 16))
        
        right_panel = ctk.CTkFrame(content_frame, fg_color="transparent")
        right_panel.pack(side="right", expand=True, fill="both", padx=(16, 0))
        
        # Left panel - File uploads
        self.create_upload_section(left_panel)
        
        # Right panel - Settings and Actions
        self.create_settings_section(right_panel)
        self.create_action_section(right_panel)
        
    def create_upload_section(self, parent):
        """Create elegant file upload section with shadow effects"""
        # Main file upload card with shadow effect (using border to simulate shadow)
        main_upload_card = ctk.CTkFrame(
            parent,
            fg_color="#ffffff",
            corner_radius=16,
            border_width=1,
            border_color="#e5e7eb"
        )
        main_upload_card.pack(fill="both", expand=True, pady=(0, 16))
        
        # Card content with padding
        card_content = ctk.CTkFrame(main_upload_card, fg_color="transparent")
        card_content.pack(fill="both", expand=True, padx=24, pady=24)
        
        # Card header
        card_header = ctk.CTkLabel(
            card_content,
            text="Main Excel File",
            font=ctk.CTkFont(family="Inter", size=18, weight="bold"),
            text_color="#0f172a"
        )
        card_header.pack(anchor="w", pady=(0, 14))
        
        # Upload button - elegant style
        self.upload_btn = ctk.CTkButton(
            card_content,
            text="Choose File",
            command=self.upload_file,
            width=240,
            height=44,
            font=ctk.CTkFont(family="Inter", size=14, weight="bold"),
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            corner_radius=10
        )
        self.upload_btn.pack(anchor="w", pady=(0, 12))
        
        # File status
        self.file_status_label = ctk.CTkLabel(
            card_content,
            text="No file selected",
            font=ctk.CTkFont(family="Inter", size=13),
            text_color="#64748b"
        )
        self.file_status_label.pack(anchor="w")
        self.file_meta_label = ctk.CTkLabel(
            card_content,
            text="Waiting for file details",
            font=ctk.CTkFont(family="Inter", size=12),
            text_color="#94a3b8"
        )
        self.file_meta_label.pack(anchor="w", pady=(6, 0))
        
        # Secondary file upload card with shadow effect
        secondary_upload_card = ctk.CTkFrame(
            parent,
            fg_color="#ffffff",
            corner_radius=16,
            border_width=1,
            border_color="#e5e7eb"
        )
        secondary_upload_card.pack(fill="both", expand=True)
        
        # Secondary card content
        secondary_content = ctk.CTkFrame(secondary_upload_card, fg_color="transparent")
        secondary_content.pack(fill="both", expand=True, padx=24, pady=24)
        
        # Secondary card header
        secondary_header = ctk.CTkLabel(
            secondary_content,
            text="Cost & MSRP Data",
            font=ctk.CTkFont(family="Inter", size=18, weight="bold"),
            text_color="#0f172a"
        )
        secondary_header.pack(anchor="w", pady=(0, 6))
        
        optional_label = ctk.CTkLabel(
            secondary_content,
            text="(Optional)",
            font=ctk.CTkFont(family="Inter", size=12),
            text_color="#94a3b8"
        )
        optional_label.pack(anchor="w", pady=(0, 16))
        
        # Secondary upload button
        self.upload_btn2 = ctk.CTkButton(
            secondary_content,
            text="Choose File",
            command=self.upload_file2,
            width=240,
            height=44,
            font=ctk.CTkFont(family="Inter", size=14, weight="bold"),
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            corner_radius=10
        )
        self.upload_btn2.pack(anchor="w", pady=(0, 12))
        
        # Secondary file status
        self.file2_status_label = ctk.CTkLabel(
            secondary_content,
            text="No file selected",
            font=ctk.CTkFont(family="Inter", size=13),
            text_color="#64748b"
        )
        self.file2_status_label.pack(anchor="w")
        self.file2_meta_label = ctk.CTkLabel(
            secondary_content,
            text="Optional file not loaded",
            font=ctk.CTkFont(family="Inter", size=12),
            text_color="#94a3b8"
        )
        self.file2_meta_label.pack(anchor="w", pady=(6, 0))
        
    def create_settings_section(self, parent):
        """Create elegant settings section with shadow effects"""
        # Settings card with shadow effect
        settings_card = ctk.CTkFrame(
            parent,
            fg_color="#ffffff",
            corner_radius=16,
            border_width=1,
            border_color="#e5e7eb"
        )
        settings_card.pack(fill="x", pady=(0, 16))
        
        # Settings content
        settings_content = ctk.CTkFrame(settings_card, fg_color="transparent")
        settings_content.pack(fill="both", padx=24, pady=24)
        
        # Settings header
        settings_header = ctk.CTkLabel(
            settings_content,
            text="Settings",
            font=ctk.CTkFont(family="Inter", size=18, weight="bold"),
            text_color="#0f172a"
        )
        settings_header.pack(anchor="w", pady=(0, 18))
        
        # Shipping cost
        shipping_container = ctk.CTkFrame(settings_content, fg_color="transparent")
        shipping_container.pack(fill="x", pady=(0, 24))
        
        shipping_label = ctk.CTkLabel(
            shipping_container,
            text="Shipping Cost",
            font=ctk.CTkFont(family="Inter", size=13, weight="bold"),
            text_color="#334155"
        )
        shipping_label.pack(anchor="w", pady=(0, 8))
        
        self.shipping_entry = ctk.CTkEntry(
            shipping_container,
            placeholder_text="0.00",
            width=260,
            height=42,
            font=ctk.CTkFont(family="Inter", size=14),
            fg_color="#f8fafc",
            border_color="#e2e8f0",
            text_color="#0f172a",
            corner_radius=10,
            border_width=1
        )
        self.shipping_entry.pack(anchor="w")
        
        # Miscellaneous cost
        misc_container = ctk.CTkFrame(settings_content, fg_color="transparent")
        misc_container.pack(fill="x", pady=(0, 24))
        
        misc_label = ctk.CTkLabel(
            misc_container,
            text="Miscellaneous Cost",
            font=ctk.CTkFont(family="Inter", size=13, weight="bold"),
            text_color="#334155"
        )
        misc_label.pack(anchor="w", pady=(0, 8))
        
        self.misc_entry = ctk.CTkEntry(
            misc_container,
            placeholder_text="0.00",
            width=260,
            height=42,
            font=ctk.CTkFont(family="Inter", size=14),
            fg_color="#f8fafc",
            border_color="#e2e8f0",
            text_color="#0f172a",
            corner_radius=10,
            border_width=1
        )
        self.misc_entry.pack(anchor="w")
        
        # Performance settings
        perf_container = ctk.CTkFrame(settings_content, fg_color="transparent")
        perf_container.pack(fill="x", pady=(0, 0))
        
        perf_label = ctk.CTkLabel(
            perf_container,
            text="Chunk Size",
            font=ctk.CTkFont(family="Inter", size=13, weight="bold"),
            text_color="#334155"
        )
        perf_label.pack(anchor="w", pady=(0, 8))
        
        self.chunk_size_var = ctk.StringVar(value="1000")
        self.chunk_size_entry = ctk.CTkEntry(
            perf_container,
            textvariable=self.chunk_size_var,
            width=260,
            height=42,
            font=ctk.CTkFont(family="Inter", size=14),
            fg_color="#f8fafc",
            border_color="#e2e8f0",
            text_color="#0f172a",
            corner_radius=10,
            border_width=1
        )
        self.chunk_size_entry.pack(anchor="w")
        
        self.chunk_hint_label = ctk.CTkLabel(
            settings_content,
            text="Suggestion updates after loading a file.",
            font=ctk.CTkFont(family="Inter", size=12),
            text_color="#94a3b8"
        )
        self.chunk_hint_label.pack(anchor="w", pady=(10, 0))
        
    def create_action_section(self, parent):
        """Create elegant action section with shadow effects"""
        # Action card with shadow effect
        action_card = ctk.CTkFrame(
            parent,
            fg_color="#ffffff",
            corner_radius=16,
            border_width=1,
            border_color="#e5e7eb"
        )
        action_card.pack(fill="x")
        
        # Action content
        action_content = ctk.CTkFrame(action_card, fg_color="transparent")
        action_content.pack(fill="both", padx=24, pady=24)
        
        # Action header
        action_header = ctk.CTkLabel(
            action_content,
            text="Process",
            font=ctk.CTkFont(family="Inter", size=18, weight="bold"),
            text_color="#0f172a"
        )
        action_header.pack(anchor="w", pady=(0, 18))
        
        # Download button - elegant and prominent
        self.download_btn = ctk.CTkButton(
            action_content,
            text="Process Excel File",
            command=self.download_file,
            state='disabled',
            width=260,
            height=50,
            font=ctk.CTkFont(family="Inter", size=15, weight="bold"),
            fg_color="#0ea5e9",
            hover_color="#0284c7",
            corner_radius=10
        )
        self.download_btn.pack(anchor="w", pady=(0, 12))
        
        # Status indicator
        self.status_label = ctk.CTkLabel(
            action_content,
            text="Ready to process (F11 toggles fullscreen, Esc exits)",
            font=ctk.CTkFont(family="Inter", size=13),
            text_color="#64748b"
        )
        self.status_label.pack(anchor="w")
        
    def create_progress_section(self):
        """Create elegant progress tracking section with shadow effects"""
        # Progress frame with shadow effect
        self.progress_frame = ctk.CTkFrame(
            self.main_container,
            fg_color="#ffffff",
            corner_radius=16,
            border_width=1,
            border_color="#e5e7eb"
        )
        self.progress_frame.pack(fill="x", padx=32, pady=(0, 32))
        
        # Progress content
        progress_content = ctk.CTkFrame(self.progress_frame, fg_color="transparent")
        progress_content.pack(fill="both", padx=24, pady=22)
        
        # Progress text
        self.progress_text = ctk.CTkLabel(
            progress_content,
            text="No file processing",
            font=ctk.CTkFont(family="Inter", size=13, weight="bold"),
            text_color="#0f172a"
        )
        self.progress_text.pack(anchor="w", pady=(0, 12))
        
        # Progress bar - elegant style
        self.progress_bar = ctk.CTkProgressBar(
            progress_content,
            width=None,
            height=10,
            corner_radius=5,
            fg_color="#e5e7eb",
            progress_color="#6366f1",
            border_width=0
        )
        self.progress_bar.pack(fill="x", pady=(0, 0))
        self.progress_bar.set(0)
        
    def apply_excel_formatting(self, save_path):
        """Apply Excel formatting to the saved file - OPTIMIZED for large files"""
        wb = load_workbook(save_path)
        ws = wb.active
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 55
        
        # Get total row count
        total_rows = ws.max_row
        print(f"Formatting {total_rows} rows with EXACT formatting for all rows...")
        self.root.after(0, lambda: self.update_progress(0.05, f"Formatting {total_rows:,} rows..."))
        format_chunk = max(500, min(5000, self.chunk_size if isinstance(self.chunk_size, int) and self.chunk_size > 0 else 1000))
        
        # Set row heights for ALL rows using memory-efficient batching
        # Process in chunks to avoid memory issues on smaller windows/machines
        for start_row in range(2, total_rows + 1, format_chunk):
            end_row = min(start_row + format_chunk - 1, total_rows)
            for row_num in range(start_row, end_row + 1):
                ws.row_dimensions[row_num].height = 50
            
            # Update progress every 10 chunks to avoid UI overload
            if (start_row // format_chunk) % 10 == 0:
                progress = 0.05 + (end_row / total_rows) * 0.10
                self.root.after(0, lambda p=progress, e=end_row: self.update_progress(p, f"Setting row heights: {e:,}/{total_rows:,}"))
                gc.collect()

        header_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        upc_col = header_map.get('UPC')
        sales_rank_col = header_map.get('Sales Rank')
        sales_30_col = header_map.get('Sales Rank 30')
        sales_90_col = header_map.get('Sales Rank 90')
        sales_180_col = header_map.get('Sales Rank 180')
        buybox_col = header_map.get('Buy Box')
        buybox_30_col = header_map.get('Buy Box 30')
        buybox_90_col = header_map.get('Buy Box 90')
        buybox_180_col = header_map.get('Buy Box 180')
        sales_badge_col = header_map.get('Sales Badge')
        amazon_col = header_map.get('Amazon')
        total_parent_ratings_col = header_map.get('Total Parent Ratings') or (ws.max_column if 'Total Parent Ratings' in [cell.value for cell in ws[1]] else None)
        total_color_ratings_col = header_map.get('Total Color Ratings') or (ws.max_column if 'Total Color Ratings' in [cell.value for cell in ws[1]] else None)
        size_col = header_map.get('Size')
        parent_col = header_map.get('Parent')
        color_col = header_map.get('Color')
        msrp_diff_col = header_map.get('MSRP Difference')
        pack_fee_col = header_map.get('Pack Fee')
        profit_col = header_map.get('Profit')
        roi_col = header_map.get('ROI')
        profit_margin_buybox_col = header_map.get('Profit Margin (Buybox)')
        profit_margin_msrp_col = header_map.get('Profit Margin (MSRP)')

        # Center, wrap, and format ALL cells - process in chunks for memory efficiency
        self.root.after(0, lambda: self.update_progress(0.10, "Aligning cells..."))
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for start_row in range(1, total_rows + 1, format_chunk):
            end_row = min(start_row + format_chunk - 1, total_rows)
            for row in ws.iter_rows(min_row=start_row, max_row=end_row):
                for cell in row:
                    cell.alignment = alignment
            
            # Update progress
            progress = 0.10 + (end_row / total_rows) * 0.05
            self.root.after(0, lambda p=progress, e=end_row: self.update_progress(p, f"Aligning cells: {e}/{total_rows}"))
            gc.collect()

        # Number formatting (comma style, no decimals) for Sales Rank columns - FORMAT ALL ROWS
        self.root.after(0, lambda: self.update_progress(0.20, "Formatting Sales Rank columns..."))
        comma_cols = [sales_rank_col, sales_30_col, sales_90_col, sales_180_col]
        for col in comma_cols:
            if col:
                for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        c.number_format = '#,##0'
                gc.collect()

        # UPC as text to prevent scientific notation - FORMAT ALL ROWS
        if upc_col:
            for cell in ws.iter_cols(min_col=upc_col, max_col=upc_col, min_row=2):
                for c in cell:
                    c.number_format = '@'  # Text format
            gc.collect()

        # Total Parent Ratings as integer (comma style) - FORMAT ALL ROWS
        if total_parent_ratings_col:
            for cell in ws.iter_cols(min_col=total_parent_ratings_col, max_col=total_parent_ratings_col, min_row=2):
                for c in cell:
                    c.number_format = '#,##0'
            gc.collect()

        # Total Color Ratings as integer (comma style) - FORMAT ALL ROWS
        if total_color_ratings_col:
            for cell in ws.iter_cols(min_col=total_color_ratings_col, max_col=total_color_ratings_col, min_row=2):
                for c in cell:
                    c.number_format = '#,##0'
            gc.collect()

        # Currency formatting for Buybox, Pack Fee, COST, and MSRP columns - FORMAT ALL ROWS
        currency_cols = [buybox_col, buybox_30_col, buybox_90_col, buybox_180_col, pack_fee_col, profit_col, header_map.get('COST'), header_map.get('MSRP')]
        for col in currency_cols:
            if col:
                for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        c.number_format = '$#,##0.00'
                gc.collect()

        # Format ROI and Profit Margin columns as numbers - FORMAT ALL ROWS
        number_cols = [roi_col, profit_margin_buybox_col, profit_margin_msrp_col]
        for col in number_cols:
            if col:
                for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        c.number_format = '#,##0.00'
                gc.collect()

        # Format AMZ In Stock % as text (since we're adding % sign manually) - FORMAT ALL ROWS
        amz_stock_col = header_map.get('AMZ In Stock %')
        if amz_stock_col:
            for cell in ws.iter_cols(min_col=amz_stock_col, max_col=amz_stock_col, min_row=2):
                for c in cell:
                    c.number_format = '@'  # Text format
            gc.collect()
        
        # Format Buy Box: % Amazon 90 days as text to preserve original format - FORMAT ALL ROWS
        buybox_amazon_col = header_map.get('Buy Box: % Amazon 90 days')
        if buybox_amazon_col:
            for cell in ws.iter_cols(min_col=buybox_amazon_col, max_col=buybox_amazon_col, min_row=2):
                for c in cell:
                    c.number_format = '@'  # Text format
            gc.collect()

        # Apply conditional formatting and colors
        self.apply_conditional_formatting(ws, header_map)
        
        # Apply thin borders to ALL cells - FORMAT ALL ROWS
        self.root.after(0, lambda: self.update_progress(0.30, "Applying borders to all cells..."))
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Process borders in chunks
        for start_row in range(1, total_rows + 1, format_chunk):
            end_row = min(start_row + format_chunk - 1, total_rows)
            for row in ws.iter_rows(min_row=start_row, max_row=end_row):
                for cell in row:
                    cell.border = border
            
            # Update progress every 10 chunks
            if (start_row // format_chunk) % 10 == 0:
                progress = 0.30 + (end_row / total_rows) * 0.20
                self.root.after(0, lambda p=progress, e=end_row: self.update_progress(p, f"Applying borders: {e:,}/{total_rows:,}"))
                gc.collect()

        # Set all column widths to exactly 15
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 15

        wb.save(save_path)

    def apply_conditional_formatting(self, ws, header_map):
        """Apply conditional formatting to ALL cells - FORMAT ALL ROWS"""
        total_rows = ws.max_row
        max_format_row = total_rows  # Format ALL rows
        print(f"Applying conditional formatting to {total_rows} rows...")
        self.root.after(0, lambda: self.update_progress(0.50, f"Applying colors: {total_rows:,} rows..."))
        
        # Color Pack Fee cells that are $7 in orange
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        pack_fee_col = header_map.get('Pack Fee')
        if pack_fee_col:
            for cell in ws.iter_cols(min_col=pack_fee_col, max_col=pack_fee_col, min_row=2, max_row=max_format_row):
                for c in cell:
                    try:
                        cell_value = float(str(c.value).replace('$', '').replace(',', ''))
                        if cell_value == 7.0:
                            c.fill = orange_fill
                    except (ValueError, TypeError, AttributeError):
                        pass

        # Conditional coloring for Sales Rank columns
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
        red_fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')

        comma_cols = [header_map.get('Sales Rank'), header_map.get('Sales Rank 30'), header_map.get('Sales Rank 90'), header_map.get('Sales Rank 180')]
        for col_idx, col in enumerate(comma_cols):
            if col:
                for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=max_format_row):
                    for c in cell:
                        try:
                            # Skip empty cells - don't color them
                            if c.value is None or c.value == '' or c.value == 0:
                                continue
                            
                            val = int(c.value)
                            if 0 < val <= 150000:
                                c.fill = green_fill
                            elif 150001 <= val <= 500000:
                                c.fill = orange_fill
                            elif val >= 500001:
                                c.fill = red_fill
                        except:
                            pass
                # Progress update after each column
                if (col_idx + 1) % 1 == 0:
                    progress = 0.50 + ((col_idx + 1) / len(comma_cols)) * 0.10
                    self.root.after(0, lambda p=progress: self.update_progress(p, f"Coloring Sales Rank columns..."))
                    gc.collect()

        # Color cells in 'Amazon Availability' column
        amazon_col = header_map.get('Amazon Availability')
        if amazon_col:
            amazon_red = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
            amazon_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            amazon_orange = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
            for row in ws.iter_rows(min_row=2, max_row=max_format_row, min_col=amazon_col, max_col=amazon_col):
                for cell in row:
                    cell_value_lower = str(cell.value).lower() if cell.value else ''
                    
                    if 'no amazon offer exists' in cell_value_lower:
                        cell.fill = amazon_green
                    elif 'amazon offer is in stock and shippable' in cell_value_lower:
                        cell.fill = amazon_red
                    elif cell.value and str(cell.value).strip():  # Any other non-empty value
                        cell.fill = amazon_orange

        # Color cells in 'Sales Badge' column if value exists
        sales_badge_col = header_map.get('Sales Badge')
        if sales_badge_col:
            badge_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            for row in ws.iter_rows(min_row=2, max_row=max_format_row, min_col=sales_badge_col, max_col=sales_badge_col):
                for cell in row:
                    if cell.value not in (None, '', 0):
                        cell.fill = badge_green

        # MSRP Difference formatting and coloring
        msrp_diff_col = header_map.get('MSRP Difference')
        if msrp_diff_col:
            for cell in ws.iter_cols(min_col=msrp_diff_col, max_col=msrp_diff_col, min_row=2, max_row=max_format_row):
                for c in cell:
                    try:
                        if isinstance(c.value, (int, float)):
                            if c.value < -0.05:
                                c.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                            elif c.value >= -0.05:
                                c.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                            c.number_format = '0.00'
                        elif c.value == 'No Buybox':
                            c.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                    except:
                        pass

        # Alternate color for Parent column only - LIMITED to first 50000 rows
        parent_col = header_map.get('Parent')
        if parent_col:
            parents = []
            for row in ws.iter_rows(min_row=2, max_row=max_format_row, min_col=parent_col, max_col=parent_col):
                for cell in row:
                    parents.append(cell.value)
            unique_parents = list(dict.fromkeys(parents))
            parent_colors = [PatternFill(start_color='FFF7CE', end_color='FFF7CE', fill_type='solid'),  # light orange
                             PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')]  # light blue
            parent_color_map = {p: parent_colors[i % 2] for i, p in enumerate(unique_parents)}
            for row in ws.iter_rows(min_row=2, max_row=max_format_row):
                parent_val = row[parent_col-1].value
                row[parent_col-1].fill = parent_color_map.get(parent_val, PatternFill())

        # Add comment to Color cell for Best Color and color it green - LIMITED to first 50000 rows
        color_col = header_map.get('Color')
        if color_col and parent_col and hasattr(self, 'best_color_map'):
            for row in ws.iter_rows(min_row=2, max_row=max_format_row):
                parent_val = row[parent_col-1].value
                color_val = row[color_col-1].value
                if self.best_color_map.get((parent_val, color_val), False):
                    row[color_col-1].comment = Comment('This color has the most ratings for this Parent ASIN', 'System')
                    row[color_col-1].fill = green_fill

        # Color COST cells that are empty/zero (no match found) - LIMITED to first 50000 rows
        cost_col = header_map.get('COST')
        if cost_col:
            for cell in ws.iter_cols(min_col=cost_col, max_col=cost_col, min_row=2, max_row=max_format_row):
                for c in cell:
                    try:
                        if c.value is None or c.value == '' or c.value == 0:
                            c.fill = red_fill
                            c.comment = Comment('No matching UPC found in cost file', 'System')
                    except:
                        pass
        
        # Color MSRP cells that are empty/None (no match found) - LIMITED to first 50000 rows
        msrp_col = header_map.get('MSRP')
        if msrp_col:
            for cell in ws.iter_cols(min_col=msrp_col, max_col=msrp_col, min_row=2, max_row=max_format_row):
                for c in cell:
                    try:
                        if c.value is None or c.value == '':
                            c.fill = red_fill
                            c.comment = Comment('No matching UPC found in cost file', 'System')
                    except:
                        pass
        
        # Color Pick & Pack cells that were originally empty (assumption = 7) - LIMITED to first 50000 rows
        pick_pack_col = header_map.get('Pick & Pack')
        if pick_pack_col:
            for cell in ws.iter_cols(min_col=pick_pack_col, max_col=pick_pack_col, min_row=2, max_row=max_format_row):
                for c in cell:
                    if isinstance(c.value, str) and '*ASSUMPTION*' in str(c.value):
                        c.value = 7.00  # Clean up the display value
                        c.fill = red_fill
                        c.comment = Comment('Assumption: Default value of 7.00 used', 'System')

        # Color Referral Fee & cells that were originally empty (assumption = 0.15) - LIMITED to first 50000 rows
        referral_fee_col = header_map.get('Referral Fee &')
        if referral_fee_col:
            for cell in ws.iter_cols(min_col=referral_fee_col, max_col=referral_fee_col, min_row=2, max_row=max_format_row):
                for c in cell:
                    if isinstance(c.value, str) and '*ASSUMPTION*' in str(c.value):
                        c.value = 0.15  # Clean up the display value
                        c.fill = red_fill
                        c.comment = Comment('Assumption: Default value of 0.15 (15%) used', 'System')

        # Conditional coloring for Profit Margin columns - LIMITED to first 50000 rows
        profit_margin_buybox_col = header_map.get('Profit Margin (Buybox)')
        profit_margin_msrp_col = header_map.get('Profit Margin (MSRP)')
        
        # Define colors for profit margin ranges
        red_fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
        orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        for col in [profit_margin_buybox_col, profit_margin_msrp_col]:
            if col:
                for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=max_format_row):
                    for c in cell:
                        try:
                            if c.value is not None and c.value != '' and c.value != 'No Buybox':
                                val = float(c.value)
                                if val < 12:
                                    c.fill = red_fill
                                elif 12 <= val <= 20:
                                    c.fill = orange_fill
                                elif val > 20:
                                    c.fill = green_fill
                            elif c.value == 'No Buybox':
                                c.fill = red_fill
                        except (ValueError, TypeError):
                            pass

    def toggle_fullscreen(self, event=None):
        """Toggle fullscreen mode"""
        if not self.is_fullscreen:
            self.root.attributes('-fullscreen', True)
            self.root.state('zoomed')
            self.is_fullscreen = True
        else:
            self.root.attributes('-fullscreen', False)
            self.root.state('normal')
            self.is_fullscreen = False
            
    def exit_fullscreen(self, event=None):
        """Exit fullscreen mode"""
        if self.is_fullscreen:
            self.root.attributes('-fullscreen', False)
            self.root.state('normal')
            self.is_fullscreen = False
        
    def upload_file(self):
        """Upload main Excel file with progress tracking"""
        file_path = filedialog.askopenfilename(
            filetypes=[('Excel Files', '*.xlsx')],
            initialdir=self.last_dir
        )
        if file_path:
            try:
                self.last_dir = os.path.dirname(file_path) or self.last_dir
                # Show loading state
                self.upload_btn.configure(text="Loading...", state="disabled")
                self.file_status_label.configure(text="Reading file...", text_color="#f59e0b")
                self.update_progress(0.1, "Reading Excel file...")
                
                # Process in background with progress tracking
                def process_file():
                    try:
                        # Read file in chunks for better memory management
                        self.df = pd.read_excel(file_path, engine='openpyxl')
                        self.file_path = file_path
                        
                        # Update UI on main thread
                        self.root.after(0, self.update_file_status_success)
                    except Exception as e:
                        error_msg = str(e)
                        self.root.after(0, lambda: self.update_file_status_error(error_msg))
                
                threading.Thread(target=process_file, daemon=True).start()
                
            except Exception as e:
                self.update_file_status_error(str(e))
                
    def update_file_status_success(self):
        """Update file status on successful upload"""
        self.upload_btn.configure(text="Choose File", state="normal")
        rows = len(self.df) if self.df is not None else 0
        cols = len(self.df.columns) if self.df is not None else 0
        file_name = os.path.basename(self.file_path)
        self.file_status_label.configure(text=f"Loaded {file_name}", text_color="#16a34a")
        self.file_meta_label.configure(text=f"{rows:,} rows x {cols:,} columns", text_color="#475569")
        self.recommend_chunk_size(rows)
        self.download_btn.configure(state='normal')
        self.update_progress(0.2, f"File loaded: {rows:,} rows")
        
    def update_file_status_error(self, error_msg):
        """Update file status on error"""
        self.upload_btn.configure(text="Choose File", state="normal")
        self.file_status_label.configure(text=f"Error: {error_msg[:50]}...", text_color="#ef4444")
        self.file_meta_label.configure(text="Fix the file and try again.", text_color="#ef4444")
        self.update_progress(0, "Error loading file")
        messagebox.showerror('Error', f'Failed to read Excel file: {error_msg}')
        
    def upload_file2(self):
        """Upload secondary file with progress tracking"""
        file_path = filedialog.askopenfilename(
            filetypes=[('Excel Files', '*.xlsx')],
            initialdir=self.last_dir
        )
        if file_path:
            try:
                self.last_dir = os.path.dirname(file_path) or self.last_dir
                # Show loading state
                self.upload_btn2.configure(text="Loading...", state="disabled")
                self.file2_status_label.configure(text="Reading file...", text_color="#f59e0b")
                
                # Process in background
                def process_file():
                    try:
                        self.df2 = pd.read_excel(file_path, engine='openpyxl')
                        self.file2_path = file_path
                        
                        # Update UI on main thread
                        self.root.after(0, self.update_file2_status_success)
                    except Exception as e:
                        error_msg = str(e)
                        self.root.after(0, lambda: self.update_file2_status_error(error_msg))
                
                threading.Thread(target=process_file, daemon=True).start()
                
            except Exception as e:
                self.update_file2_status_error(str(e))
                
    def update_file2_status_success(self):
        """Update secondary file status on successful upload"""
        self.upload_btn2.configure(text="Choose File", state="normal")
        file_name = os.path.basename(self.file2_path)
        rows = len(self.df2) if self.df2 is not None else 0
        cols = len(self.df2.columns) if self.df2 is not None else 0
        self.file2_status_label.configure(text=f"Loaded {file_name}", text_color="#16a34a")
        self.file2_meta_label.configure(text=f"{rows:,} rows x {cols:,} columns", text_color="#475569")
        
    def update_file2_status_error(self, error_msg):
        """Update secondary file status on error"""
        self.upload_btn2.configure(text="Choose File", state="normal")
        self.file2_status_label.configure(text=f"Error: {error_msg[:50]}...", text_color="#ef4444")
        self.file2_meta_label.configure(text="Optional file failed to load.", text_color="#ef4444")
        messagebox.showerror('Error', f'Failed to read Cost/MSRP file: {error_msg}')
        
    def recommend_chunk_size(self, row_count):
        """Suggest a chunk size based on detected row count to balance speed and memory."""
        default_chunk = 1000
        try:
            rows = int(row_count)
        except (TypeError, ValueError):
            self.chunk_hint_label.configure(text="Using default chunk size: 1,000 rows.")
            return default_chunk
        
        recommended = max(500, min(4000, max(500, rows // 4)))
        current = (self.chunk_size_var.get() or "").strip()
        if current in ("", "1000"):
            self.chunk_size_var.set(str(recommended))
        self.chunk_hint_label.configure(
            text=f"Suggested chunk size: {recommended:,} rows (based on {rows:,} rows detected)."
        )
        return recommended
        
    def update_progress(self, value, text):
        """Update progress bar and text"""
        self.progress_bar.set(value)
        self.progress_text.configure(text=text)

    def download_file(self):
        """Download formatted file with progress tracking"""
        if self.df is None:
            messagebox.showerror('Error', 'No file uploaded.')
            return
            
        if self.processing:
            messagebox.showwarning('Warning', 'File is already being processed.')
            return
            
        file_path_str = str(self.file_path)
        if not file_path_str or not os.path.isfile(file_path_str):
            messagebox.showerror('Error', f'Invalid file path: {file_path_str}')
            return
            
        base, ext = os.path.splitext(file_path_str)
        save_path = base + '_formatted.xlsx'
        
        # Update chunk size from settings
        try:
            user_chunk = int((self.chunk_size_var.get() or "0").replace(",", ""))
            self.chunk_size = max(200, min(5000, user_chunk))
        except:
            self.chunk_size = 1000
        self.chunk_size_var.set(str(self.chunk_size))
        self.chunk_hint_label.configure(text=f"Using chunk size: {self.chunk_size:,} rows per batch.")
        
        # Show processing state
        self.processing = True
        self.download_btn.configure(text="Processing...", state="disabled")
        self.status_label.configure(text="Processing file...", text_color="#f59e0b")
        self.update_progress(0.1, "Starting processing...")
        
        # Process in background with progress tracking
        def process_and_save():
            try:
                self.format_and_save_excel_optimized(save_path)
                self.root.after(0, lambda: self.update_download_success(save_path))
            except Exception as e:
                error_msg = str(e)
                self.root.after(0, lambda: self.update_download_error(error_msg))
                
        threading.Thread(target=process_and_save, daemon=True).start()

    def update_download_success(self, save_path):
        """Update UI on successful download"""
        self.processing = False
        self.download_btn.configure(text="Process Excel File", state="normal")
        self.status_label.configure(text="File processed successfully", text_color="#10b981")
        self.update_progress(1.0, f"File saved: {os.path.basename(save_path)}")
        messagebox.showinfo('Success', f'File saved to {save_path}\n\nThe formatted Excel file will open automatically!')
        
    def update_download_error(self, error_msg):
        """Update UI on download error"""
        self.processing = False
        self.download_btn.configure(text="Process Excel File", state="normal")
        self.status_label.configure(text="Processing failed", text_color="#ef4444")
        self.update_progress(0, "Error occurred")
        messagebox.showerror('Error', f'Failed to save file: {error_msg}')

    def format_and_save_excel_optimized(self, save_path):
        """Optimized Excel processing for large datasets"""
        try:
            # Update progress
            self.root.after(0, lambda: self.update_progress(0.2, "Preparing data..."))
            
            # Create a copy of the dataframe
            df = self.df.copy()
            
            # Delete Locale and Image columns if they exist
            columns_to_delete = ['Locale', 'Image']
            for col in columns_to_delete:
                if col in df.columns:
                    df.drop(columns=[col], inplace=True)
            
            # Rename columns
            df.rename(columns=HEADER_MAP, inplace=True)
            self.root.after(0, lambda: self.update_progress(0.3, "Processing columns..."))
            
            # Handle Pack Fee - only fill empty cells with 7 and mark for special formatting
            if 'Pick & Pack' in df.columns:
                # Mark which cells were originally empty
                empty_mask = df['Pick & Pack'].isna() | (df['Pick & Pack'] == '') | (df['Pick & Pack'].astype(str) == '0')
                df['Pick & Pack'] = pd.to_numeric(df['Pick & Pack'], errors='coerce')
                df['Pick & Pack'] = df['Pick & Pack'].fillna(7.0)
                # Round to 2 decimal places
                df['Pick & Pack'] = df['Pick & Pack'].round(2)
                # Add special marker for originally empty cells
                df.loc[empty_mask, 'Pick & Pack'] = '7.00*ASSUMPTION*'
            
            # Process UPC, EAN, and GTIN columns - clean and format
            if 'UPC' in df.columns:
                self.root.after(0, lambda: self.update_progress(0.35, "Processing product codes..."))
                
                # Clean UPC column - remove .0 from the end of numeric strings
                df['UPC'] = df['UPC'].astype(str)
                df['UPC'] = df['UPC'].replace(['nan', 'NaN', 'None'], '')
                df['UPC'] = df['UPC'].str.replace(r'\.0$', '', regex=True)
                
                print(f"DEBUG: Processing {len(df)} rows with UPC codes")
                print(f"DEBUG: Sample UPC values: {df['UPC'].head().tolist()}")
            
            # Clean EAN column if it exists
            if 'EAN' in df.columns:
                df['EAN'] = df['EAN'].astype(str)
                df['EAN'] = df['EAN'].replace(['nan', 'NaN', 'None'], '')
                df['EAN'] = df['EAN'].str.replace(r'\.0$', '', regex=True)
                print(f"DEBUG: Sample EAN values: {df['EAN'].head().tolist()}")
            
            # Clean GTIN column if it exists
            if 'GTIN' in df.columns:
                df['GTIN'] = df['GTIN'].astype(str)
                df['GTIN'] = df['GTIN'].replace(['nan', 'NaN', 'None'], '')
                df['GTIN'] = df['GTIN'].str.replace(r'\.0$', '', regex=True)
                print(f"DEBUG: Sample GTIN values: {df['GTIN'].head().tolist()}")
            
            # Clean Imported by Code column if it exists
            if 'Imported by Code' in df.columns:
                df['Imported by Code'] = df['Imported by Code'].astype(str)
                df['Imported by Code'] = df['Imported by Code'].replace(['nan', 'NaN', 'None'], '')
                df['Imported by Code'] = df['Imported by Code'].str.replace(r'\.0$', '', regex=True)
                print(f"DEBUG: Processing {len(df)} rows with Imported by Code")
                print(f"DEBUG: Sample Imported by Code values: {df['Imported by Code'].head().tolist()}")
                print(f"DEBUG: Imported by Code column exists: {('Imported by Code' in df.columns)}")
            
            # Merge Cost/MSRP if second file is uploaded
            if self.df2 is not None:
                self.root.after(0, lambda: self.update_progress(0.4, "Merging cost data..."))
                df2 = self.df2.copy()
                df2.columns = [str(col).strip() for col in df2.columns]
                
                # Look for matching code column in cost file - try "Imported by Code" first, then "UPC" as fallback
                imported_code_col2 = None
                cost_col2 = None
                msrp_col2 = None
                for col in df2.columns:
                    col_upper = col.upper()
                    if 'IMPORTED BY CODE' in col_upper or ('IMPORTED' in col_upper and 'CODE' in col_upper):
                        imported_code_col2 = col
                        print(f"DEBUG: Found Imported by Code column in cost file: '{col}'")
                    elif 'UPC' in col_upper and imported_code_col2 is None:
                        # Fallback to UPC if Imported by Code not found
                        imported_code_col2 = col
                        print(f"DEBUG: Found UPC column in cost file (using as Imported by Code): '{col}'")
                    if 'COST' in col_upper and cost_col2 is None:
                        cost_col2 = col
                        print(f"DEBUG: Found COST column in cost file: '{col}'")
                    if 'MSRP' in col_upper and msrp_col2 is None:
                        msrp_col2 = col
                        print(f"DEBUG: Found MSRP column in cost file: '{col}'")
                
                print(f"DEBUG: Cost file columns: {list(df2.columns)}")
                print(f"DEBUG: Matching code column: {imported_code_col2}, COST column: {cost_col2}, MSRP column: {msrp_col2}")
                
                if imported_code_col2 and (cost_col2 or msrp_col2):
                    merge_cols = [imported_code_col2]
                    if cost_col2: merge_cols.append(cost_col2)
                    if msrp_col2: merge_cols.append(msrp_col2)
                    df2 = df2[merge_cols]
                    df2.rename(columns={imported_code_col2: 'Imported by Code'}, inplace=True)
                    
                    # Normalize "Imported by Code" for matching - handle both UPC and other codes
                    def normalize_imported_code(code):
                        """Normalize imported code for matching"""
                        if pd.isna(code) or code == '':
                            return None
                        
                        # Convert to string first to preserve leading zeros
                        code_str = str(code).strip()
                        
                        # Remove .0 from end (Excel sometimes adds this)
                        if code_str.endswith('.0'):
                            code_str = code_str[:-2]
                        
                        # Return None if empty or invalid
                        if not code_str or code_str.lower() in ['nan', 'none', '']:
                            return None
                        
                        # If it's a numeric code (likely UPC), normalize it
                        # Remove common separators for numeric codes
                        if code_str.replace('-', '').replace(' ', '').replace('_', '').isdigit():
                            # Remove separators
                            code_str = code_str.replace('-', '').replace(' ', '').replace('_', '')
                            # Pad with leading zeros if it's a short UPC (UPCs are typically 12 digits)
                            if len(code_str) < 12 and len(code_str) >= 8:
                                code_str = code_str.zfill(12)
                        
                        # Return the cleaned code
                        return code_str
                    
                    # Create normalized merge columns for cost file
                    df2['merge_code'] = df2['Imported by Code'].apply(normalize_imported_code)
                    
                    # Create set of all valid codes from cost file for matching
                    cost_code_set = set(df2['merge_code'].dropna().unique())
                    print(f"DEBUG: Cost file has {len(cost_code_set)} unique Imported by Code values")
                    
                    # Check if main file has "Imported by Code" column, or fallback to UPC/EAN/GTIN
                    main_code_col = None
                    if 'Imported by Code' in df.columns:
                        main_code_col = 'Imported by Code'
                        print(f"DEBUG: Found 'Imported by Code' column in main file with {df['Imported by Code'].notna().sum()} non-null values")
                    elif 'UPC' in df.columns:
                        main_code_col = 'UPC'
                        print(f"DEBUG: Using 'UPC' column from main file as fallback (found {df['UPC'].notna().sum()} non-null values)")
                    elif 'EAN' in df.columns:
                        main_code_col = 'EAN'
                        print(f"DEBUG: Using 'EAN' column from main file as fallback (found {df['EAN'].notna().sum()} non-null values)")
                    elif 'GTIN' in df.columns:
                        main_code_col = 'GTIN'
                        print(f"DEBUG: Using 'GTIN' column from main file as fallback (found {df['GTIN'].notna().sum()} non-null values)")
                    
                    if main_code_col is None:
                        print("WARNING: No matching code column found in main file.")
                        print(f"DEBUG: Available columns in main file: {list(df.columns)[:20]}...")
                        print("WARNING: Matching will be skipped. Please ensure your main file has an 'Imported by Code' column (or UPC/EAN/GTIN as fallback).")
                    else:
                        # Normalize codes in main file
                        df['merge_code'] = df[main_code_col].apply(normalize_imported_code)
                        
                        # Remove rows where merge_code is None from cost file
                        df2_merge = df2[df2['merge_code'].notna()].copy()
                        
                        print(f"DEBUG: Main file has {len(df)} rows, cost file has {len(df2_merge)} valid code rows")
                        print(f"DEBUG: Sample merge codes from main file: {df['merge_code'].dropna().head(10).tolist()}")
                        print(f"DEBUG: Sample merge codes from cost file: {df2_merge['merge_code'].head(10).tolist()}")
                        
                        # Show which codes will match
                        main_codes = set(df['merge_code'].dropna().unique())
                        cost_codes = set(df2_merge['merge_code'].dropna().unique())
                        matching_codes = main_codes & cost_codes
                        print(f"DEBUG: Found {len(matching_codes)} matching codes out of {len(main_codes)} unique codes in main file")
                        print(f"DEBUG: Sample matching codes: {list(matching_codes)[:10]}")
                        
                        # Perform the merge directly on the main dataframe
                        df = pd.merge(df, df2_merge, left_on='merge_code', right_on='merge_code', how='left', suffixes=('', '_cost'))
                        
                        # Rename cost and MSRP columns to standard names after merge
                        if cost_col2 and cost_col2 in df.columns:
                            df.rename(columns={cost_col2: 'COST'}, inplace=True)
                        if msrp_col2 and msrp_col2 in df.columns:
                            df.rename(columns={msrp_col2: 'MSRP'}, inplace=True)
                        
                        # Count successful matches
                        matches_found = df['COST'].notna().sum() if 'COST' in df.columns else (df['MSRP'].notna().sum() if 'MSRP' in df.columns else 0)
                        print(f"DEBUG: Successfully merged {matches_found} rows with cost/MSRP data")
                        
                        # Clean up the merge columns and any duplicate columns ending with _cost
                        df.drop(columns=['merge_code'], errors='ignore', inplace=True)
                        
                        # Remove any columns ending with '_cost' (duplicates from merge)
                        cost_suffix_cols = [col for col in df.columns if col.endswith('_cost')]
                        if cost_suffix_cols:
                            df.drop(columns=cost_suffix_cols, errors='ignore', inplace=True)
                            print(f"DEBUG: Removed duplicate columns: {cost_suffix_cols}")
                    
                    # Ensure 'Cost' is numeric and positive (handle both 'Cost' and 'COST')
                    cost_col_name = 'COST' if 'COST' in df.columns else 'Cost'
                    if cost_col_name in df.columns:
                        df[cost_col_name] = pd.to_numeric(df[cost_col_name], errors='coerce')
                        df[cost_col_name] = df[cost_col_name].apply(lambda x: x if pd.notna(x) and x > 0 else 0)
                        print(f"DEBUG: Found {(df[cost_col_name] > 0).sum()} rows with valid COST data")
                    
                    # Ensure 'MSRP' is numeric, positive, and correctly scaled
                    if 'MSRP' in df.columns:
                        df['MSRP'] = pd.to_numeric(df['MSRP'], errors='coerce')
                        df['MSRP'] = df['MSRP'].apply(lambda x: x if pd.notna(x) and x > 0 else None)
                        print(f"DEBUG: Found {df['MSRP'].notna().sum()} rows with valid MSRP data")

            # Remove unnamed columns
            df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
            
            # Remove rows that are problematic (empty, mostly empty, or have key fields as 'nan')
            def is_problematic_row(row):
                # Check if row is completely empty
                if row.isna().all():
                    return True
                
                # Check if all values are empty strings
                empty_string_count = sum(1 for val in row if str(val).strip() == '')
                if empty_string_count == len(row):
                    return True
                
                # Special case: If Imported by Code is 'nan' and we have minimal other data, remove the row
                if 'Imported by Code' in row.index:
                    imported_code_value = str(row['Imported by Code']).strip().lower()
                    if imported_code_value == 'nan' or imported_code_value == '':
                        # Count non-empty, non-zero values (excluding Imported by Code)
                        non_empty_count = 0
                        for field in row.index:
                            if field != 'Imported by Code':
                                val_str = str(row[field]).strip().lower()
                                if val_str not in ['nan', '', 'none', '0', '0.0', '0.0%']:
                                    non_empty_count += 1
                        
                        # If we have 3 or fewer meaningful values (excluding Imported by Code), remove the row
                        if non_empty_count <= 3:
                            return True
                
                # Also check UPC for backward compatibility (if Imported by Code doesn't exist)
                if 'Imported by Code' not in row.index and 'UPC' in row.index:
                    upc_value = str(row['UPC']).strip().lower()
                    if upc_value == 'nan':
                        # Count non-empty, non-zero values (excluding UPC)
                        non_empty_count = 0
                        for field in row.index:
                            if field != 'UPC':
                                val_str = str(row[field]).strip().lower()
                                if val_str not in ['nan', '', 'none', '0', '0.0', '0.0%']:
                                    non_empty_count += 1
                        
                        # If we have 3 or fewer meaningful values (excluding UPC), remove the row
                        if non_empty_count <= 3:
                            return True
                
                # Check if key identifying fields are 'nan' or empty
                key_fields = ['Brand', 'Parent', 'ASIN', 'Title']
                for field in key_fields:
                    if field in row.index:
                        value = str(row[field]).strip().lower()
                        if value in ['nan', '', 'none']:
                            # If any key field is nan/empty, check if the row has minimal data
                            non_empty_count = sum(1 for val in row.values if str(val).strip().lower() not in ['nan', '', 'none'])
                            if non_empty_count <= 2:  # Allow rows with 2 or fewer non-empty values
                                return True
                            break
                
                return False
            
            mask = df.apply(is_problematic_row, axis=1)
            df = df[~mask]
            
            self.root.after(0, lambda: self.update_progress(0.5, "Cleaning data..."))

            # Add costs
            shipping_cost = 0.0
            misc_cost = 0.0
            try:
                shipping_cost = float(self.shipping_entry.get()) if self.shipping_entry.get() else 0.0
            except:
                shipping_cost = 0.0
            try:
                misc_cost = float(self.misc_entry.get()) if self.misc_entry.get() else 0.0
            except:
                misc_cost = 0.0
                
            # Add shipping and misc costs to the cost column
            cost_col_name = 'COST' if 'COST' in df.columns else 'Cost'
            if cost_col_name in df.columns:
                df[cost_col_name] = pd.to_numeric(df[cost_col_name], errors='coerce').fillna(0) + shipping_cost + misc_cost

            # Process data in chunks for better performance
            self.root.after(0, lambda: self.update_progress(0.6, "Calculating metrics..."))
            
            # Add calculated columns
            # Fix: Convert categorical columns to string to avoid pandas category ordering issues
            if 'Parent' in df.columns and 'Ratings' in df.columns:
                # Convert Parent to string to avoid category ordering issues
                df['Parent'] = df['Parent'].astype(str)
                df['Ratings'] = pd.to_numeric(df['Ratings'], errors='coerce').fillna(0).astype(int)
                parent_ratings_sum = df.groupby('Parent', observed=True)['Ratings'].transform('sum')
                df['Total Parent Ratings'] = parent_ratings_sum
                # Release memory
                gc.collect()

            if 'Parent' in df.columns and 'Color' in df.columns and 'Ratings' in df.columns:
                # Convert Color to string to avoid category ordering issues
                df['Color'] = df['Color'].astype(str)
                color_ratings_sum = df.groupby(['Parent', 'Color'], observed=True)['Ratings'].transform('sum')
                df['Total Color Ratings'] = color_ratings_sum
                # Release memory
                gc.collect()

            # Identify best color(s) for each Parent based on Rating - Child
            best_color_map = {}
            if 'Parent' in df.columns and 'Color' in df.columns and 'Rating - Child' in df.columns:
                # Convert Rating - Child to numeric to handle string values
                df['Rating - Child'] = pd.to_numeric(df['Rating - Child'], errors='coerce').fillna(0)
                # Process in chunks to avoid memory issues with large datasets
                chunk_size = min(50000, len(df))
                for i in range(0, len(df), chunk_size):
                    chunk = df.iloc[i:i+chunk_size]
                    for parent, group in chunk.groupby('Parent', observed=True):
                        max_ratings = group['Rating - Child'].max()
                        best_colors = group[group['Rating - Child'] == max_ratings]['Color'].unique()
                        for color in best_colors:
                            best_color_map[(parent, color)] = True
                    # Release memory after each chunk
                    del chunk
                    gc.collect()
            
            # Store the best color map for later use in formatting
            self.best_color_map = best_color_map

            # Calculate Total Ratings Color - sum of Rating - Child for each color of each Parent ASIN
            if 'Parent' in df.columns and 'Color' in df.columns and 'Rating - Child' in df.columns:
                # Convert Rating - Child to numeric to handle string values
                df['Rating - Child'] = pd.to_numeric(df['Rating - Child'], errors='coerce').fillna(0)
                
                # Group by Parent and Color, then sum Rating - Child
                total_ratings_by_color = df.groupby(['Parent', 'Color'], observed=True)['Rating - Child'].sum().reset_index()
                total_ratings_by_color.rename(columns={'Rating - Child': 'Total Ratings Color'}, inplace=True)
                
                # Merge back to main dataframe
                df = pd.merge(df, total_ratings_by_color, on=['Parent', 'Color'], how='left', suffixes=('', '_sum'))
                
                # Fill any NaN values with 0
                df['Total Ratings Color'] = df['Total Ratings Color'].fillna(0)
                
                print(f"DEBUG: Calculated Total Ratings Color for {len(total_ratings_by_color)} Parent-Color combinations")

            # Handle Referral Fee & - fill empty cells with 0.15 and mark for special formatting
            if 'Referral Fee &' in df.columns:
                # Mark which cells were originally empty
                empty_mask = df['Referral Fee &'].isna() | (df['Referral Fee &'] == '') | (df['Referral Fee &'].astype(str) == '0')
                df['Referral Fee &'] = pd.to_numeric(df['Referral Fee &'], errors='coerce')
                df['Referral Fee &'] = df['Referral Fee &'].fillna(0.15)
                # Round to 2 decimal places
                df['Referral Fee &'] = df['Referral Fee &'].round(2)
                # Add special marker for originally empty cells
                df.loc[empty_mask, 'Referral Fee &'] = '0.15*ASSUMPTION*'

            # Sort data - FIX: Convert all sort columns to string to avoid category ordering issues
            sort_cols = [col for col in ['Parent', 'Color', 'Size'] if col in df.columns]
            if sort_cols:
                # Convert all sort columns to string first
                for col in sort_cols:
                    if col in df.columns:
                        df[col] = df[col].astype(str)
                # Now sort
                df.sort_values(by=sort_cols, inplace=True)

            # Keep AMZ In Stock % in original format (no conversion or % sign addition)
            if 'AMZ In Stock %' in df.columns:
                # Keep original values as-is, just ensure they're properly formatted as text
                # Replace 'nan' with empty string to keep blank cells blank
                df['AMZ In Stock %'] = df['AMZ In Stock %'].astype(str)
                df['AMZ In Stock %'] = df['AMZ In Stock %'].replace(['nan', 'NaN', 'None'], '')
            
            # Keep Buy Box: % Amazon 90 days in original text format (no conversion)
            if 'Buy Box: % Amazon 90 days' in df.columns:
                # Keep original values as-is, preserve text format with % signs
                df['Buy Box: % Amazon 90 days'] = df['Buy Box: % Amazon 90 days'].astype(str)
                df['Buy Box: % Amazon 90 days'] = df['Buy Box: % Amazon 90 days'].replace(['nan', 'NaN', 'None'], '')

            # Add calculated columns - OPTIMIZED for large datasets
            self.root.after(0, lambda: self.update_progress(0.7, "Calculating profits..."))
            
            # Process in chunks to prevent memory issues with large datasets
            chunk_size = min(50000, len(df))
            results = []
            
            for i in range(0, len(df), chunk_size):
                chunk = df.iloc[i:i+chunk_size].copy()
                
                # Calculate all metrics for this chunk
                chunk['Profit'] = chunk.apply(self.calc_profit, axis=1)
                chunk['ROI'] = chunk.apply(self.calc_roi, axis=1)
                chunk['Profit Margin (Buybox)'] = chunk.apply(self.calc_profit_margin_buybox, axis=1)
                chunk['Profit Margin (MSRP)'] = chunk.apply(self.calc_profit_margin_msrp, axis=1)
                chunk['MSRP Difference'] = chunk.apply(self.msrp_diff, axis=1)
                
                results.append(chunk)
                
                # Release memory after each chunk
                del chunk
                gc.collect()
                
                # Update progress
                progress = 0.7 + (min(i + chunk_size, len(df)) / len(df)) * 0.1
                self.root.after(0, lambda p=progress: self.update_progress(p, f"Processing rows {i+1:,} to {min(i+chunk_size, len(df)):,}..."))
            
            # Combine all chunks back into dataframe
            df = pd.concat(results, ignore_index=True)
            
            # Release memory
            del results
            gc.collect()

            # Drop unnamed columns
            if 'Unnamed: 19' in df.columns:
                df.drop(columns=['Unnamed: 19'], inplace=True)

            self.root.after(0, lambda: self.update_progress(0.8, "Saving to Excel..."))
            
            # Release memory before saving
            gc.collect()
            
            # Save to Excel with optimized settings for large files
            df.to_excel(save_path, index=False, engine='openpyxl')
            
            # Release memory after saving
            del df
            gc.collect()
            
            # Apply formatting
            self.root.after(0, lambda: self.update_progress(0.9, "Applying formatting..."))
            self.apply_excel_formatting(save_path)
            
            # Final memory cleanup
            gc.collect()
            
            # Auto-open the file
            self.root.after(0, lambda: self.update_progress(1.0, "Opening file..."))
            self.auto_open_excel(save_path)
            
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")

    def clean_price(self, val):
        """Clean price values - handle strings, floats, and edge cases"""
        if val is None or val == '' or val == 0:
            return None
        
        if isinstance(val, str):
            # Remove currency symbols and whitespace
            val = val.replace('$', '').replace(',', '').strip()
            if val == '' or val == '0':
                return None
        
        try:
            cleaned_val = float(val)
            return cleaned_val if cleaned_val > 0 else None
        except (ValueError, TypeError):
            return None

    def msrp_diff(self, row):
        """
        Calculate MSRP Difference as a number: Buy Box - MSRP (not percentage)
        """
        msrp = self.clean_price(row.get('MSRP', None))
        buybox = self.clean_price(row.get('Buy Box', None))
        buybox_30 = self.clean_price(row.get('Buy Box 30', None))
        buybox_90 = self.clean_price(row.get('Buy Box 90', None))
        buybox_180 = self.clean_price(row.get('Buy Box 180', None))
        buybox_val = None
        for val in [buybox, buybox_30, buybox_90, buybox_180]:
            if val is not None and val > 0:
                buybox_val = val
                break
        if msrp is not None and msrp > 0 and buybox_val is not None and buybox_val > 0:
            diff = buybox_val - msrp
            return diff
        elif msrp is not None and msrp > 0 and buybox_val is None:
            return 'No Buybox'
        else:
            return ''

    def calc_profit(self, row):
        cost = 0
        try:
            cost = float(row.get('COST', 0))
        except (ValueError, TypeError):
            cost = 0
            
        # Handle Pack Fee (formerly FBA Pick&Pack Fee)
        pack_fee = 7.0  # Default value
        try:
            pack_fee_val = row.get('Pick & Pack')
            if pd.notna(pack_fee_val) and pack_fee_val != '':
                # Handle assumption marker
                if isinstance(pack_fee_val, str) and '*ASSUMPTION*' in pack_fee_val:
                    pack_fee = 7.0
                else:
                    pack_fee = float(pack_fee_val)
        except (ValueError, TypeError):
            pass
            
        # Add Pack Fee to total cost
        cost += pack_fee

        # Get referral fee percentage, default to 15% if empty
        referral_fee_pct = 0.15  # Default 15%
        try:
            ref_fee = row.get('Referral Fee &')
            if pd.notna(ref_fee) and str(ref_fee).strip() != '':
                # Handle assumption marker
                if isinstance(ref_fee, str) and '*ASSUMPTION*' in ref_fee:
                    referral_fee_pct = 0.15
                else:
                    # Handle both decimal (0.17) and percentage ("17.00%") formats
                    ref_fee = str(ref_fee).replace('%', '').strip()
                    if float(ref_fee) > 1:  # If number is like 17.00
                        referral_fee_pct = float(ref_fee) / 100
                    else:  # If number is like 0.17
                        referral_fee_pct = float(ref_fee)
        except (ValueError, TypeError, AttributeError):
            pass
            
        for col in ['Buy Box', 'Buy Box 30', 'Buy Box 90', 'Buy Box 180', 'MSRP']:
            val = self.clean_price(row.get(col, None))
            if val is not None and val > 0:
                # Calculate revenue after referral fee
                revenue = val * (1 - referral_fee_pct)
                return round(revenue - cost, 2)  # Round to 2 decimal places
        return -cost if cost else ''

    def calc_roi(self, row):
        """Calculate Return on Investment (ROI) as a percentage"""
        profit = row.get('Profit', 0)
        if isinstance(profit, str):
            return ''
        cost = 0
        try:
            cost = float(row.get('COST', 0))
        except (ValueError, TypeError):
            cost = 0
        if cost > 0 and isinstance(profit, (int, float)):
            roi = (profit / cost) * 100
            return roi if roi != float('inf') else ''
        return ''

    def calc_profit_margin_buybox(self, row):
        """Calculate Profit Margin as a percentage based on Buy Box with fallback logic"""
        # Try Buy Box first, then Buy Box 30, then Buy Box 90, then Buy Box 180
        buybox_val = None
        for col in ['Buy Box', 'Buy Box 30', 'Buy Box 90', 'Buy Box 180']:
            try:
                val = self.clean_price(row.get(col, None))
                if val is not None and val > 0:
                    buybox_val = val
                    break
            except (ValueError, TypeError):
                continue
        
        if buybox_val is None:
            return 'No Buybox'
        
        # Get referral fee percentage
        referral_fee_pct = 0.15  # Default 15%
        try:
            ref_fee = row.get('Referral Fee &')
            if pd.notna(ref_fee) and str(ref_fee).strip() != '':
                # Handle assumption marker
                if isinstance(ref_fee, str) and '*ASSUMPTION*' in ref_fee:
                    referral_fee_pct = 0.15
                else:
                    ref_fee = str(ref_fee).replace('%', '').strip()
                    if float(ref_fee) > 1:  # If number is like 17.00
                        referral_fee_pct = float(ref_fee) / 100
                    else:  # If number is like 0.17
                        referral_fee_pct = float(ref_fee)
        except (ValueError, TypeError, AttributeError):
            pass
        
        # Calculate revenue after referral fee
        revenue = buybox_val * (1 - referral_fee_pct)
        
        # Get total cost (COST + Pick & Pack Fee)
        cost = 0
        try:
            cost = float(row.get('COST', 0))
        except (ValueError, TypeError):
            cost = 0
        
        # Add Pick & Pack Fee
        pack_fee = 7.0  # Default value
        try:
            pack_fee_val = row.get('Pick & Pack')
            if pd.notna(pack_fee_val) and pack_fee_val != '':
                # Handle assumption marker
                if isinstance(pack_fee_val, str) and '*ASSUMPTION*' in pack_fee_val:
                    pack_fee = 7.0
                else:
                    pack_fee = float(pack_fee_val)
        except (ValueError, TypeError):
            pass
        
        total_cost = cost + pack_fee
        
        # Calculate profit margin: (Profit / Sale Price) × 100
        # Profit = Revenue - Total Cost, Sale Price = Buy Box Price
        if buybox_val > 0:
            profit = revenue - total_cost
            margin = (profit / buybox_val) * 100
            return round(margin, 2) if margin != float('inf') else ''
        return ''

    def calc_profit_margin_msrp(self, row):
        """Calculate Profit Margin as a percentage based on MSRP from input file"""
        # Use MSRP from the input UPC-COST-MSRP file as the revenue base
        msrp = self.clean_price(row.get('MSRP', None))
        if msrp is None or msrp <= 0:
            return ''
        
        # Get referral fee percentage
        referral_fee_pct = 0.15  # Default 15%
        try:
            ref_fee = row.get('Referral Fee &')
            if pd.notna(ref_fee) and str(ref_fee).strip() != '':
                # Handle assumption marker
                if isinstance(ref_fee, str) and '*ASSUMPTION*' in ref_fee:
                    referral_fee_pct = 0.15
                else:
                    ref_fee = str(ref_fee).replace('%', '').strip()
                    if float(ref_fee) > 1:  # If number is like 17.00
                        referral_fee_pct = float(ref_fee) / 100
                    else:  # If number is like 0.17
                        referral_fee_pct = float(ref_fee)
        except (ValueError, TypeError, AttributeError):
            pass
        
        # Calculate revenue after referral fee
        revenue = msrp * (1 - referral_fee_pct)
        
        # Get total cost (COST + Pick & Pack Fee)
        cost = 0
        try:
            cost = float(row.get('COST', 0))
        except (ValueError, TypeError):
            cost = 0
        
        # Add Pick & Pack Fee
        pack_fee = 7.0  # Default value
        try:
            pack_fee_val = row.get('Pick & Pack')
            if pd.notna(pack_fee_val) and pack_fee_val != '':
                # Handle assumption marker
                if isinstance(pack_fee_val, str) and '*ASSUMPTION*' in pack_fee_val:
                    pack_fee = 7.0
                else:
                    pack_fee = float(pack_fee_val)
        except (ValueError, TypeError):
            pass
        
        total_cost = cost + pack_fee
        
        # Calculate profit margin: (Profit / Sale Price) × 100
        # Profit = Revenue - Total Cost, Sale Price = MSRP
        if msrp > 0:
            profit = revenue - total_cost
            margin = (profit / msrp) * 100
            return round(margin, 2) if margin != float('inf') else ''
        return ''


    def auto_open_excel(self, file_path):
        """Automatically open the Excel file with the default application"""
        try:
            if platform.system() == "Windows":
                os.startfile(file_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", file_path], check=True)
            else:  # Linux
                subprocess.run(["xdg-open", file_path], check=True)
        except Exception as e:
            print(f"Could not auto-open file: {e}")

if __name__ == '__main__':
    root = ctk.CTk()
    app = ExcelFormatterApp(root)
    root.mainloop()
