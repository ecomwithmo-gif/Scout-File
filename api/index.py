from flask import Flask, request, jsonify, send_file, render_template_string
import pandas as pd
import os
import tempfile
import gc
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from werkzeug.utils import secure_filename
import json

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

# Header mapping as per requirements
HEADER_MAP = {
    'Brand': 'Brand',
    'Parent ASIN': 'Parent',
    'ASIN': 'ASIN',
    'Product Codes: UPC': 'UPC',
    'Product Codes: EAN': 'EAN',
    'Product Codes: GTIN': 'GTIN',
    'Imported by Code': 'Imported by Code',
    'Title': 'Title',
    'Color': 'Color',
    'Size': 'Size',
    'Bought in past month': 'Sales Badge',
    'Reviews: Rating Count': 'Rating Count',
    'Reviews: Review Count - Format Specific': 'Rating - Child',
    'Sales Rank: Current': 'Sales Rank',
    'Sales Rank: 30 days avg.': 'Sales Rank 30',
    'Sales Rank: 90 days avg.': 'Sales Rank 90',
    'Sales Rank: 180 days avg.': 'Sales Rank 180',
    'Buy Box 🚚: Current': 'Buy Box',
    'Buy Box 🚚: 30 days avg.': 'Buy Box 30',
    'Buy Box 🚚: 90 days avg.': 'Buy Box 90',
    'Buy Box 🚚: 180 days avg.': 'Buy Box 180',
    'Amazon: 90 days OOS': 'AMZ In Stock %',
    'Buy Box: % Amazon 90 days': 'Buy Box: % Amazon 90 days',
    'Amazon: Availability of the Amazon offer': 'Amazon Availability',
    'Count of retrieved live offers: New, FBA': 'FBA',
    'Count of retrieved live offers: New, FBM': 'FBM',
    'FBA Pick&Pack Fee': 'Pick & Pack',
    'Referral Fee %': 'Referral Fee &',
}

# HTML Template
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Formatter Pro - Web App</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            padding: 40px;
        }
        
        h1 {
            color: #1a202c;
            font-size: 2.5rem;
            margin-bottom: 10px;
            text-align: center;
        }
        
        .subtitle {
            color: #718096;
            font-size: 1.2rem;
            margin-bottom: 30px;
            text-align: center;
        }
        
        .upload-section {
            background: #f7fafc;
            border: 2px dashed #cbd5e0;
            border-radius: 12px;
            padding: 40px;
            text-align: center;
            margin-bottom: 30px;
            transition: all 0.3s;
        }
        
        .upload-section.dragover {
            border-color: #667eea;
            background: #edf2f7;
        }
        
        .file-input-wrapper {
            position: relative;
            display: inline-block;
        }
        
        .file-input {
            position: absolute;
            opacity: 0;
            width: 100%;
            height: 100%;
            cursor: pointer;
        }
        
        .file-label {
            display: inline-block;
            padding: 15px 40px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 10px;
            cursor: pointer;
            font-weight: bold;
            font-size: 1.1rem;
            transition: transform 0.2s;
        }
        
        .file-label:hover {
            transform: translateY(-2px);
        }
        
        .file-info {
            margin-top: 20px;
            color: #4a5568;
        }
        
        .settings-section {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }
        
        .setting-group {
            display: flex;
            flex-direction: column;
        }
        
        .setting-group label {
            color: #2d3748;
            font-weight: bold;
            margin-bottom: 8px;
        }
        
        .setting-group input {
            padding: 12px;
            border: 2px solid #e2e8f0;
            border-radius: 8px;
            font-size: 1rem;
            transition: border-color 0.2s;
        }
        
        .setting-group input:focus {
            outline: none;
            border-color: #667eea;
        }
        
        .process-btn {
            width: 100%;
            padding: 18px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 1.2rem;
            font-weight: bold;
            cursor: pointer;
            transition: transform 0.2s;
            margin-bottom: 20px;
        }
        
        .process-btn:hover:not(:disabled) {
            transform: translateY(-2px);
        }
        
        .process-btn:disabled {
            opacity: 0.6;
            cursor: not-allowed;
        }
        
        .progress-section {
            margin-top: 30px;
            display: none;
        }
        
        .progress-section.active {
            display: block;
        }
        
        .progress-bar {
            width: 100%;
            height: 30px;
            background: #e2e8f0;
            border-radius: 15px;
            overflow: hidden;
            margin-bottom: 10px;
        }
        
        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
            width: 0%;
            transition: width 0.3s;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
        }
        
        .progress-text {
            text-align: center;
            color: #4a5568;
            margin-top: 10px;
        }
        
        .download-section {
            margin-top: 30px;
            text-align: center;
            display: none;
        }
        
        .download-section.active {
            display: block;
        }
        
        .download-btn {
            display: inline-block;
            padding: 15px 40px;
            background: #10b981;
            color: white;
            border-radius: 10px;
            text-decoration: none;
            font-weight: bold;
            font-size: 1.1rem;
            transition: transform 0.2s;
        }
        
        .download-btn:hover {
            transform: translateY(-2px);
        }
        
        .error-message {
            background: #fee;
            color: #c33;
            padding: 15px;
            border-radius: 8px;
            margin-top: 20px;
            display: none;
        }
        
        .error-message.active {
            display: block;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Excel Formatter Pro</h1>
        <p class="subtitle">Advanced Excel Processing Web Application</p>
        
        <form id="uploadForm" enctype="multipart/form-data">
            <div class="upload-section" id="uploadSection">
                <div class="file-input-wrapper">
                    <input type="file" id="mainFile" name="main_file" accept=".xlsx" class="file-input" required>
                    <label for="mainFile" class="file-label">Choose Main Excel File</label>
                </div>
                <div class="file-info" id="mainFileInfo">No file selected</div>
            </div>
            
            <div class="upload-section">
                <div class="file-input-wrapper">
                    <input type="file" id="costFile" name="cost_file" accept=".xlsx" class="file-input">
                    <label for="costFile" class="file-label">Choose Cost & MSRP File (Optional)</label>
                </div>
                <div class="file-info" id="costFileInfo">No file selected (optional)</div>
            </div>
            
            <div class="settings-section">
                <div class="setting-group">
                    <label for="shippingCost">Shipping Cost</label>
                    <input type="number" id="shippingCost" name="shipping_cost" step="0.01" value="0.00" placeholder="0.00">
                </div>
                <div class="setting-group">
                    <label for="miscCost">Miscellaneous Cost</label>
                    <input type="number" id="miscCost" name="misc_cost" step="0.01" value="0.00" placeholder="0.00">
                </div>
                <div class="setting-group">
                    <label for="chunkSize">Chunk Size</label>
                    <input type="number" id="chunkSize" name="chunk_size" value="1000" placeholder="1000">
                </div>
            </div>
            
            <button type="submit" class="process-btn" id="processBtn">Process Excel File</button>
        </form>
        
        <div class="progress-section" id="progressSection">
            <div class="progress-bar">
                <div class="progress-fill" id="progressFill">0%</div>
            </div>
            <div class="progress-text" id="progressText">Processing...</div>
        </div>
        
        <div class="download-section" id="downloadSection">
            <a href="#" class="download-btn" id="downloadBtn">Download Formatted File</a>
        </div>
        
        <div class="error-message" id="errorMessage"></div>
    </div>
    
    <script>
        const uploadSection = document.getElementById('uploadSection');
        const mainFileInput = document.getElementById('mainFile');
        const costFileInput = document.getElementById('costFile');
        const mainFileInfo = document.getElementById('mainFileInfo');
        const costFileInfo = document.getElementById('costFileInfo');
        const uploadForm = document.getElementById('uploadForm');
        const processBtn = document.getElementById('processBtn');
        const progressSection = document.getElementById('progressSection');
        const progressFill = document.getElementById('progressFill');
        const progressText = document.getElementById('progressText');
        const downloadSection = document.getElementById('downloadSection');
        const downloadBtn = document.getElementById('downloadBtn');
        const errorMessage = document.getElementById('errorMessage');
        
        // Drag and drop handlers
        uploadSection.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadSection.classList.add('dragover');
        });
        
        uploadSection.addEventListener('dragleave', () => {
            uploadSection.classList.remove('dragover');
        });
        
        uploadSection.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadSection.classList.remove('dragover');
            if (e.dataTransfer.files.length > 0) {
                mainFileInput.files = e.dataTransfer.files;
                updateFileInfo(mainFileInput, mainFileInfo);
            }
        });
        
        mainFileInput.addEventListener('change', () => updateFileInfo(mainFileInput, mainFileInfo));
        costFileInput.addEventListener('change', () => updateFileInfo(costFileInput, costFileInfo));
        
        function updateFileInfo(input, infoElement) {
            if (input.files.length > 0) {
                const file = input.files[0];
                const size = (file.size / (1024 * 1024)).toFixed(2);
                infoElement.textContent = `${file.name} (${size} MB)`;
            } else {
                infoElement.textContent = input.id === 'mainFile' ? 'No file selected' : 'No file selected (optional)';
            }
        }
        
        uploadForm.addEventListener('submit', async (e) => {
            e.preventDefault();
            
            const formData = new FormData(uploadForm);
            
            // Reset UI
            processBtn.disabled = true;
            processBtn.textContent = 'Processing...';
            progressSection.classList.add('active');
            downloadSection.classList.remove('active');
            errorMessage.classList.remove('active');
            
            try {
                const response = await fetch('/api/process', {
                    method: 'POST',
                    body: formData
                });
                
                if (!response.ok) {
                    const error = await response.json();
                    throw new Error(error.error || 'Processing failed');
                }
                
                const result = await response.json();
                
                if (result.success) {
                    progressFill.style.width = '100%';
                    progressFill.textContent = '100%';
                    progressText.textContent = 'Processing complete!';
                    
                    downloadBtn.href = result.download_url;
                    downloadSection.classList.add('active');
                } else {
                    throw new Error(result.error || 'Processing failed');
                }
            } catch (error) {
                errorMessage.textContent = error.message;
                errorMessage.classList.add('active');
                progressSection.classList.remove('active');
            } finally {
                processBtn.disabled = false;
                processBtn.textContent = 'Process Excel File';
            }
        });
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    return HTML_TEMPLATE

@app.route('/api/process', methods=['POST'])
def process_file():
    try:
        # Get uploaded files
        main_file = request.files.get('main_file')
        cost_file = request.files.get('cost_file')
        
        if not main_file:
            return jsonify({'success': False, 'error': 'No main file uploaded'}), 400
        
        # Get settings
        shipping_cost = float(request.form.get('shipping_cost', 0) or 0)
        misc_cost = float(request.form.get('misc_cost', 0) or 0)
        chunk_size = int(request.form.get('chunk_size', 1000) or 1000)
        
        # Save uploaded files temporarily
        main_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(main_file.filename))
        main_file.save(main_path)
        
        cost_path = None
        if cost_file and cost_file.filename:
            cost_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(cost_file.filename))
            cost_file.save(cost_path)
        
        # Process the Excel file
        output_path = process_excel_file(main_path, cost_path, shipping_cost, misc_cost, chunk_size)
        
        # Generate download URL
        filename = os.path.basename(output_path)
        download_url = f'/api/download/{filename}'
        
        return jsonify({
            'success': True,
            'download_url': download_url,
            'filename': filename
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(filename))
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(file_path, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def process_excel_file(main_path, cost_path, shipping_cost, misc_cost, chunk_size):
    """Process Excel file with the same logic as desktop app"""
    # Read main file
    df = pd.read_excel(main_path, engine='openpyxl')
    
    # Delete Locale and Image columns if they exist
    columns_to_delete = ['Locale', 'Image']
    for col in columns_to_delete:
        if col in df.columns:
            df.drop(columns=[col], inplace=True)
    
    # Rename columns
    df.rename(columns=HEADER_MAP, inplace=True)
    
    # Handle Pack Fee
    if 'Pick & Pack' in df.columns:
        empty_mask = df['Pick & Pack'].isna() | (df['Pick & Pack'] == '') | (df['Pick & Pack'].astype(str) == '0')
        df['Pick & Pack'] = pd.to_numeric(df['Pick & Pack'], errors='coerce')
        df['Pick & Pack'] = df['Pick & Pack'].fillna(7.0).round(2)
        df.loc[empty_mask, 'Pick & Pack'] = '7.00*ASSUMPTION*'
    
    # Clean UPC, EAN, GTIN columns
    for col in ['UPC', 'EAN', 'GTIN', 'Imported by Code']:
        if col in df.columns:
            df[col] = df[col].astype(str)
            df[col] = df[col].replace(['nan', 'NaN', 'None'], '')
            df[col] = df[col].str.replace(r'\.0$', '', regex=True)
    
    # Merge Cost/MSRP if second file is uploaded
    if cost_path:
        df2 = pd.read_excel(cost_path, engine='openpyxl')
        df2.columns = [str(col).strip() for col in df2.columns]
        
        # Find matching columns
        imported_code_col2 = None
        cost_col2 = None
        msrp_col2 = None
        
        for col in df2.columns:
            col_upper = col.upper()
            if 'IMPORTED BY CODE' in col_upper or ('IMPORTED' in col_upper and 'CODE' in col_upper):
                imported_code_col2 = col
            elif 'UPC' in col_upper and imported_code_col2 is None:
                imported_code_col2 = col
            if 'COST' in col_upper and cost_col2 is None:
                cost_col2 = col
            if 'MSRP' in col_upper and msrp_col2 is None:
                msrp_col2 = col
        
        if imported_code_col2 and (cost_col2 or msrp_col2):
            merge_cols = [imported_code_col2]
            if cost_col2:
                merge_cols.append(cost_col2)
            if msrp_col2:
                merge_cols.append(msrp_col2)
            
            df2 = df2[merge_cols]
            df2.rename(columns={imported_code_col2: 'Imported by Code'}, inplace=True)
            
            # Normalize codes for matching
            def normalize_code(code):
                if pd.isna(code) or code == '':
                    return None
                code_str = str(code).strip()
                if code_str.endswith('.0'):
                    code_str = code_str[:-2]
                if not code_str or code_str.lower() in ['nan', 'none', '']:
                    return None
                if code_str.replace('-', '').replace(' ', '').replace('_', '').isdigit():
                    code_str = code_str.replace('-', '').replace(' ', '').replace('_', '')
                    if len(code_str) < 12 and len(code_str) >= 8:
                        code_str = code_str.zfill(12)
                return code_str
            
            df2['merge_code'] = df2['Imported by Code'].apply(normalize_code)
            
            # Determine main file code column
            main_code_col = None
            if 'Imported by Code' in df.columns:
                main_code_col = 'Imported by Code'
            elif 'UPC' in df.columns:
                main_code_col = 'UPC'
            elif 'EAN' in df.columns:
                main_code_col = 'EAN'
            elif 'GTIN' in df.columns:
                main_code_col = 'GTIN'
            
            if main_code_col:
                df['merge_code'] = df[main_code_col].apply(normalize_code)
                df2_merge = df2[df2['merge_code'].notna()].copy()
                
                df = pd.merge(df, df2_merge, left_on='merge_code', right_on='merge_code', how='left', suffixes=('', '_cost'))
                
                if cost_col2 and cost_col2 in df.columns:
                    df.rename(columns={cost_col2: 'COST'}, inplace=True)
                if msrp_col2 and msrp_col2 in df.columns:
                    df.rename(columns={msrp_col2: 'MSRP'}, inplace=True)
                
                df.drop(columns=['merge_code'], errors='ignore', inplace=True)
                cost_suffix_cols = [col for col in df.columns if col.endswith('_cost')]
                if cost_suffix_cols:
                    df.drop(columns=cost_suffix_cols, errors='ignore', inplace=True)
            
            # Ensure COST and MSRP are numeric
            cost_col_name = 'COST' if 'COST' in df.columns else 'Cost'
            if cost_col_name in df.columns:
                df[cost_col_name] = pd.to_numeric(df[cost_col_name], errors='coerce')
                df[cost_col_name] = df[cost_col_name].apply(lambda x: x if pd.notna(x) and x > 0 else 0)
            
            if 'MSRP' in df.columns:
                df['MSRP'] = pd.to_numeric(df['MSRP'], errors='coerce')
                df['MSRP'] = df['MSRP'].apply(lambda x: x if pd.notna(x) and x > 0 else None)
    
    # Remove unnamed columns
    df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
    
    # Add shipping and misc costs
    cost_col_name = 'COST' if 'COST' in df.columns else 'Cost'
    if cost_col_name in df.columns:
        df[cost_col_name] = pd.to_numeric(df[cost_col_name], errors='coerce').fillna(0) + shipping_cost + misc_cost
    
    # Calculate metrics
    # Handle Referral Fee
    if 'Referral Fee &' in df.columns:
        empty_mask = df['Referral Fee &'].isna() | (df['Referral Fee &'] == '') | (df['Referral Fee &'].astype(str) == '0')
        df['Referral Fee &'] = pd.to_numeric(df['Referral Fee &'], errors='coerce')
        df['Referral Fee &'] = df['Referral Fee &'].fillna(0.15).round(2)
        df.loc[empty_mask, 'Referral Fee &'] = '0.15*ASSUMPTION*'
    
    # Calculate Total Parent Ratings and Total Color Ratings
    if 'Parent' in df.columns and 'Rating Count' in df.columns:
        df['Parent'] = df['Parent'].astype(str)
        df['Rating Count'] = pd.to_numeric(df['Rating Count'], errors='coerce').fillna(0).astype(int)
        parent_ratings_sum = df.groupby('Parent', observed=True)['Rating Count'].transform('sum')
        df['Total Parent Ratings'] = parent_ratings_sum
        gc.collect()
    
    if 'Parent' in df.columns and 'Color' in df.columns and 'Rating Count' in df.columns:
        df['Color'] = df['Color'].astype(str)
        color_ratings_sum = df.groupby(['Parent', 'Color'], observed=True)['Rating Count'].transform('sum')
        df['Total Color Ratings'] = color_ratings_sum
        gc.collect()
    
    # Calculate Total Ratings Color
    if 'Parent' in df.columns and 'Color' in df.columns and 'Rating - Child' in df.columns:
        df['Rating - Child'] = pd.to_numeric(df['Rating - Child'], errors='coerce').fillna(0)
        total_ratings_by_color = df.groupby(['Parent', 'Color'], observed=True)['Rating - Child'].sum().reset_index()
        total_ratings_by_color.rename(columns={'Rating - Child': 'Total Ratings Color'}, inplace=True)
        df = pd.merge(df, total_ratings_by_color, on=['Parent', 'Color'], how='left', suffixes=('', '_sum'))
        df['Total Ratings Color'] = df['Total Ratings Color'].fillna(0)
    
    # Keep AMZ In Stock % and Buy Box: % Amazon 90 days in original format
    if 'AMZ In Stock %' in df.columns:
        df['AMZ In Stock %'] = df['AMZ In Stock %'].astype(str)
        df['AMZ In Stock %'] = df['AMZ In Stock %'].replace(['nan', 'NaN', 'None'], '')
    
    if 'Buy Box: % Amazon 90 days' in df.columns:
        df['Buy Box: % Amazon 90 days'] = df['Buy Box: % Amazon 90 days'].astype(str)
        df['Buy Box: % Amazon 90 days'] = df['Buy Box: % Amazon 90 days'].replace(['nan', 'NaN', 'None'], '')
    
    # Add calculated columns
    df['Profit'] = df.apply(calc_profit, axis=1)
    df['ROI'] = df.apply(calc_roi, axis=1)
    df['Profit Margin (Buybox)'] = df.apply(calc_profit_margin_buybox, axis=1)
    df['Profit Margin (MSRP)'] = df.apply(calc_profit_margin_msrp, axis=1)
    df['MSRP Difference'] = df.apply(msrp_diff, axis=1)
    
    # Sort data
    sort_cols = [col for col in ['Parent', 'Color', 'Size'] if col in df.columns]
    if sort_cols:
        for col in sort_cols:
            if col in df.columns:
                df[col] = df[col].astype(str)
        df.sort_values(by=sort_cols, inplace=True)
    
    # Save to Excel
    base_name = os.path.splitext(os.path.basename(main_path))[0]
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{base_name}_formatted.xlsx')
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Apply formatting
    apply_excel_formatting(output_path, chunk_size)
    
    return output_path

def clean_price(val):
    """Clean price values - handle strings, floats, and edge cases"""
    if val is None or val == '' or val == 0:
        return None
    
    if isinstance(val, str):
        val = val.replace('$', '').replace(',', '').strip()
        if val == '' or val == '0':
            return None
    
    try:
        cleaned_val = float(val)
        return cleaned_val if cleaned_val > 0 else None
    except (ValueError, TypeError):
        return None

def msrp_diff(row):
    """Calculate MSRP Difference as a number: Buy Box - MSRP"""
    msrp = clean_price(row.get('MSRP', None))
    buybox = clean_price(row.get('Buy Box', None))
    buybox_30 = clean_price(row.get('Buy Box 30', None))
    buybox_90 = clean_price(row.get('Buy Box 90', None))
    buybox_180 = clean_price(row.get('Buy Box 180', None))
    buybox_val = None
    for val in [buybox, buybox_30, buybox_90, buybox_180]:
        if val is not None and val > 0:
            buybox_val = val
            break
    if msrp is not None and msrp > 0 and buybox_val is not None and buybox_val > 0:
        return buybox_val - msrp
    elif msrp is not None and msrp > 0 and buybox_val is None:
        return 'No Buybox'
    else:
        return ''

def calc_profit(row):
    cost = 0
    try:
        cost = float(row.get('COST', 0))
    except (ValueError, TypeError):
        cost = 0
    
    pack_fee = 7.0
    try:
        pack_fee_val = row.get('Pick & Pack')
        if pd.notna(pack_fee_val) and pack_fee_val != '':
            if isinstance(pack_fee_val, str) and '*ASSUMPTION*' in pack_fee_val:
                pack_fee = 7.0
            else:
                pack_fee = float(pack_fee_val)
    except (ValueError, TypeError):
        pass
    
    cost += pack_fee
    
    referral_fee_pct = 0.15
    try:
        ref_fee = row.get('Referral Fee &')
        if pd.notna(ref_fee) and str(ref_fee).strip() != '':
            if isinstance(ref_fee, str) and '*ASSUMPTION*' in ref_fee:
                referral_fee_pct = 0.15
            else:
                ref_fee = str(ref_fee).replace('%', '').strip()
                if float(ref_fee) > 1:
                    referral_fee_pct = float(ref_fee) / 100
                else:
                    referral_fee_pct = float(ref_fee)
    except (ValueError, TypeError, AttributeError):
        pass
    
    for col in ['Buy Box', 'Buy Box 30', 'Buy Box 90', 'Buy Box 180', 'MSRP']:
        val = clean_price(row.get(col, None))
        if val is not None and val > 0:
            revenue = val * (1 - referral_fee_pct)
            return round(revenue - cost, 2)
    return -cost if cost else ''

def calc_roi(row):
    """Calculate Return on Investment (ROI) as a percentage"""
    profit = row.get('Profit', 0)
    if isinstance(profit, str):
        return ''
    cost = 0
    try:
        cost = float(row.get('COST', 0))
    except (ValueError, TypeError):
        cost = 0
    if cost > 0 and isinstance(profit, (int, float)):
        roi = (profit / cost) * 100
        return roi if roi != float('inf') else ''
    return ''

def calc_profit_margin_buybox(row):
    """Calculate Profit Margin as a percentage based on Buy Box"""
    buybox_val = None
    for col in ['Buy Box', 'Buy Box 30', 'Buy Box 90', 'Buy Box 180']:
        try:
            val = clean_price(row.get(col, None))
            if val is not None and val > 0:
                buybox_val = val
                break
        except (ValueError, TypeError):
            continue
    
    if buybox_val is None:
        return 'No Buybox'
    
    referral_fee_pct = 0.15
    try:
        ref_fee = row.get('Referral Fee &')
        if pd.notna(ref_fee) and str(ref_fee).strip() != '':
            if isinstance(ref_fee, str) and '*ASSUMPTION*' in ref_fee:
                referral_fee_pct = 0.15
            else:
                ref_fee = str(ref_fee).replace('%', '').strip()
                if float(ref_fee) > 1:
                    referral_fee_pct = float(ref_fee) / 100
                else:
                    referral_fee_pct = float(ref_fee)
    except (ValueError, TypeError, AttributeError):
        pass
    
    revenue = buybox_val * (1 - referral_fee_pct)
    
    cost = 0
    try:
        cost = float(row.get('COST', 0))
    except (ValueError, TypeError):
        cost = 0
    
    pack_fee = 7.0
    try:
        pack_fee_val = row.get('Pick & Pack')
        if pd.notna(pack_fee_val) and pack_fee_val != '':
            if isinstance(pack_fee_val, str) and '*ASSUMPTION*' in pack_fee_val:
                pack_fee = 7.0
            else:
                pack_fee = float(pack_fee_val)
    except (ValueError, TypeError):
        pass
    
    total_cost = cost + pack_fee
    
    if buybox_val > 0:
        profit = revenue - total_cost
        margin = (profit / buybox_val) * 100
        return round(margin, 2) if margin != float('inf') else ''
    return ''

def calc_profit_margin_msrp(row):
    """Calculate Profit Margin as a percentage based on MSRP"""
    msrp = clean_price(row.get('MSRP', None))
    if msrp is None or msrp <= 0:
        return ''
    
    referral_fee_pct = 0.15
    try:
        ref_fee = row.get('Referral Fee &')
        if pd.notna(ref_fee) and str(ref_fee).strip() != '':
            if isinstance(ref_fee, str) and '*ASSUMPTION*' in ref_fee:
                referral_fee_pct = 0.15
            else:
                ref_fee = str(ref_fee).replace('%', '').strip()
                if float(ref_fee) > 1:
                    referral_fee_pct = float(ref_fee) / 100
                else:
                    referral_fee_pct = float(ref_fee)
    except (ValueError, TypeError, AttributeError):
        pass
    
    revenue = msrp * (1 - referral_fee_pct)
    
    cost = 0
    try:
        cost = float(row.get('COST', 0))
    except (ValueError, TypeError):
        cost = 0
    
    pack_fee = 7.0
    try:
        pack_fee_val = row.get('Pick & Pack')
        if pd.notna(pack_fee_val) and pack_fee_val != '':
            if isinstance(pack_fee_val, str) and '*ASSUMPTION*' in pack_fee_val:
                pack_fee = 7.0
            else:
                pack_fee = float(pack_fee_val)
    except (ValueError, TypeError):
        pass
    
    total_cost = cost + pack_fee
    
    if msrp > 0:
        profit = revenue - total_cost
        margin = (profit / msrp) * 100
        return round(margin, 2) if margin != float('inf') else ''
    return ''

def apply_excel_formatting(file_path, chunk_size):
    """Apply Excel formatting with conditional formatting"""
    wb = load_workbook(file_path)
    ws = wb.active
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 55
    
    total_rows = ws.max_row
    format_chunk = max(500, min(5000, chunk_size))
    
    # Set row heights
    for row_num in range(2, total_rows + 1):
        ws.row_dimensions[row_num].height = 50
    
    # Get header map
    header_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
    
    # Alignment
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
    
    # Number formatting - Sales Rank columns
    sales_rank_cols = [
        header_map.get('Sales Rank'),
        header_map.get('Sales Rank 30'),
        header_map.get('Sales Rank 90'),
        header_map.get('Sales Rank 180')
    ]
    for col in sales_rank_cols:
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '#,##0'
    
    # UPC as text
    upc_col = header_map.get('UPC')
    if upc_col:
        for cell in ws.iter_cols(min_col=upc_col, max_col=upc_col, min_row=2):
            for c in cell:
                c.number_format = '@'
    
    # Total Parent/Color Ratings as integers
    for col_name in ['Total Parent Ratings', 'Total Color Ratings']:
        col = header_map.get(col_name)
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '#,##0'
    
    # Currency formatting
    currency_cols = [
        header_map.get('Buy Box'),
        header_map.get('Buy Box 30'),
        header_map.get('Buy Box 90'),
        header_map.get('Buy Box 180'),
        header_map.get('Pick & Pack'),
        header_map.get('Profit'),
        header_map.get('COST'),
        header_map.get('MSRP')
    ]
    for col in currency_cols:
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '$#,##0.00'
    
    # ROI and Profit Margin formatting
    number_cols = [
        header_map.get('ROI'),
        header_map.get('Profit Margin (Buybox)'),
        header_map.get('Profit Margin (MSRP)')
    ]
    for col in number_cols:
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '#,##0.00'
    
    # Text formatting for percentage columns
    text_cols = [
        header_map.get('AMZ In Stock %'),
        header_map.get('Buy Box: % Amazon 90 days')
    ]
    for col in text_cols:
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '@'
    
    # Conditional formatting - Sales Rank colors
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
    
    for col in sales_rank_cols:
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    try:
                        if c.value is None or c.value == '' or c.value == 0:
                            continue
                        val = int(c.value)
                        if 0 < val <= 150000:
                            c.fill = green_fill
                        elif 150001 <= val <= 500000:
                            c.fill = orange_fill
                        elif val >= 500001:
                            c.fill = red_fill
                    except:
                        pass
    
    # Conditional formatting - Pack Fee (orange for $7)
    pack_fee_col = header_map.get('Pick & Pack')
    if pack_fee_col:
        for cell in ws.iter_cols(min_col=pack_fee_col, max_col=pack_fee_col, min_row=2):
            for c in cell:
                try:
                    cell_value = float(str(c.value).replace('$', '').replace(',', ''))
                    if cell_value == 7.0:
                        c.fill = orange_fill
                except:
                    pass
    
    # Conditional formatting - Amazon Availability
    amazon_col = header_map.get('Amazon Availability')
    if amazon_col:
        amazon_red = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
        amazon_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        amazon_orange = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
        for row in ws.iter_rows(min_row=2, min_col=amazon_col, max_col=amazon_col):
            for cell in row:
                cell_value_lower = str(cell.value).lower() if cell.value else ''
                if 'no amazon offer exists' in cell_value_lower:
                    cell.fill = amazon_green
                elif 'amazon offer is in stock and shippable' in cell_value_lower:
                    cell.fill = amazon_red
                elif cell.value and str(cell.value).strip():
                    cell.fill = amazon_orange
    
    # Conditional formatting - Sales Badge
    sales_badge_col = header_map.get('Sales Badge')
    if sales_badge_col:
        badge_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        for row in ws.iter_rows(min_row=2, min_col=sales_badge_col, max_col=sales_badge_col):
            for cell in row:
                if cell.value not in (None, '', 0):
                    cell.fill = badge_green
    
    # Conditional formatting - MSRP Difference
    msrp_diff_col = header_map.get('MSRP Difference')
    if msrp_diff_col:
        for cell in ws.iter_cols(min_col=msrp_diff_col, max_col=msrp_diff_col, min_row=2):
            for c in cell:
                try:
                    if isinstance(c.value, (int, float)):
                        if c.value < -0.05:
                            c.fill = red_fill
                        elif c.value >= -0.05:
                            c.fill = green_fill
                        c.number_format = '0.00'
                    elif c.value == 'No Buybox':
                        c.fill = red_fill
                except:
                    pass
    
    # Conditional formatting - Profit Margin
    for col_name in ['Profit Margin (Buybox)', 'Profit Margin (MSRP)']:
        col = header_map.get(col_name)
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    try:
                        if c.value is not None and c.value != '' and c.value != 'No Buybox':
                            val = float(c.value)
                            if val < 12:
                                c.fill = red_fill
                            elif 12 <= val <= 20:
                                c.fill = orange_fill
                            elif val > 20:
                                c.fill = green_fill
                        elif c.value == 'No Buybox':
                            c.fill = red_fill
                    except:
                        pass
    
    # Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    
    # Column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 15
    
    wb.save(file_path)

if __name__ == '__main__':
    app.run(debug=True)
